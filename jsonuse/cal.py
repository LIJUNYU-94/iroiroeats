import openpyxl
from openpyxl.styles import Alignment

# Excelファイルのパス
file_path = "keisu.xlsx"  # 必ず正しいファイル名に変更
output_path = "keisucal.xlsx"  # 出力ファイル

# openpyxlでExcelを開く
wb = openpyxl.load_workbook(file_path)
ws = wb.active  # 最初のシートを取得

# **Excelの列番号を定義**
column_map = {"X": 24, "Y": 25, "Z": 26, "AA": 27, "AB": 28}
new_column_map = {"AH": 34, "AI": 35, "AJ": 36, "AK": 37, "AL": 38}  # 変換後の列番号

# **スケーリング係数（事前に計算済み）**
scaling_factors = {
    "X": 0.078,  # エネルギー
    "Y": 2.02,   # たん白質
    "Z": 2.44,   # 脂質
    "AA": 0.12,  # ミネラル
    "AB": 0.16   # ビタミン
}

# **結合セルの情報を取得**
merge_ranges = list(ws.merged_cells.ranges)

# **結合セルごとに処理**
for merge_range in merge_ranges:
    if merge_range.min_col == 24:  # 'X' 列を基準（ExcelではX=24列目）
        start_row = merge_range.min_row
        end_row = merge_range.max_row

        # 各列ごとに処理
        for key, col_idx in column_map.items():
            new_col_idx = new_column_map[key.replace("X", "AH").replace("Y", "AI").replace("Z", "AJ").replace("AA", "AK").replace("AB", "AL")]

            # **結合セルの一番上の値を取得し、スケーリング**
            cell_value = ws.cell(row=start_row, column=col_idx).value
            if cell_value is None:
                cell_value = 0  # データがない場合は0にする

            if isinstance(cell_value, (int, float)):  # 数値のみを対象
                scaled_value = cell_value * scaling_factors[key]  # 係数をかける
                ws.cell(row=start_row, column=new_col_idx, value=scaled_value)  # AH, AI, AJ, AK, AL の上部セルに書き込む
wb.save(output_path)
print(f"処理完了！出力ファイル: {output_path}")