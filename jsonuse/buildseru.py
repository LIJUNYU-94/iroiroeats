import openpyxl
from openpyxl.styles import Alignment

# **Excelファイルのパス**
file_path = "data.xlsx"  # 処理するExcelファイル
output_path = "output.xlsx"  # 出力ファイル

# **openpyxlでExcelを開く**
wb = openpyxl.load_workbook(file_path)
ws = wb.active  # 最初のシートを取得

# **EF列の結合セルを取得**
merge_ranges = list(ws.merged_cells.ranges)


# **結合セルごとにX, Y, Z, AA, AB列のデータをSUM関数で計算して格納**
for merge_range in merge_ranges:
    if merge_range.min_col == 5:  # 'E'列を基準にする（EFの左側）
        start_row = merge_range.min_row
        end_row = merge_range.max_row

        # **SUM関数を作成**
        sum_energy = f"=SUM(I{start_row}:I{end_row})"
        sum_protein = f"=SUM(J{start_row}:J{end_row})"
        sum_fat = f"=SUM(K{start_row}:K{end_row})"
        sum_mineral = f"=SUM(V{start_row}:V{end_row})"
        sum_vitamin = f"=SUM(W{start_row}:W{end_row})"

        # **計算結果をX, Y, Z, AA, AB列にセット**
        ws[f"X{start_row}"] = sum_energy
        ws[f"Y{start_row}"] = sum_protein
        ws[f"Z{start_row}"] = sum_fat
        ws[f"AA{start_row}"] = sum_mineral
        ws[f"AB{start_row}"] = sum_vitamin

        # **セル結合**
        ws.merge_cells(f"X{start_row}:X{end_row}")
        ws.merge_cells(f"Y{start_row}:Y{end_row}")
        ws.merge_cells(f"Z{start_row}:Z{end_row}")
        ws.merge_cells(f"AA{start_row}:AA{end_row}")
        ws.merge_cells(f"AB{start_row}:AB{end_row}")

        # **セルの中央揃え**
        for col in ["X", "Y", "Z", "AA", "AB"]:
            ws[f"{col}{start_row}"].alignment = Alignment(horizontal="center", vertical="center")

# **Excelファイルを保存**
wb.save(output_path)
print(f"処理完了！出力ファイル: {output_path}")
